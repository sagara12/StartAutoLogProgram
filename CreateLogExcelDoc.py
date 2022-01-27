from datetime import timedelta, timezone
from datetime import datetime as dt
import datetime
from CreateDatePath import CreateDatePath
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl import styles
import win32com.client as win32
import win32com
from pathlib import Path

class CreateLogExcelDoc:

    def create_excel_doc_data(self, infor_dic):

        # 딕셔너리에서 해당 리스트 추출
        colum_list = infor_dic['colum_List']
        data_list = infor_dic['search_List']

        # total_List
        total_List = []

        # for문을 돌면서 data_list에 있는 정보 추출 및 수정
        for i in range(len(data_list)):

            #수정한 데이터를 저장할 리스트 선언
            value_List = []

            #data_list의 i 번째 데이터를 저장할 리스트 선언
            row_list = []
            row_list = data_list[i]

            # data_list에서 데이터 추출
            User = row_list[0]
            Mashine = row_list[1]
            Group_val = row_list[2]
            session_day_start = row_list[3]
            session_day_end = row_list[4]
            reconnect_count = row_list[5]
            session_day_start_result = session_day_start + timedelta(hours=9) # 한국 시간은 UTC 기준 +9시간 이기 때문에 9시간을 더 해줌

            #value_List에 데이터 주입
            value_List.append(User)
            value_List.append(Mashine)
            value_List.append(Group_val)
            value_List.append(session_day_start_result)

            if (str(type(session_day_end)) != "<class 'NoneType'>"):  # session이 종료 안되어 있을경우 session_day_end 칼럼에 None이 들어 가기 때문에 타입 체크 필수

                # session_day_end 칼럼이 None이 아닐경우
                session_day_end_result = session_day_end + timedelta(hours=9) # 한국 시간은 UTC 기준 +9시간 이기 때문에 9시간을 더 해줌

                # value_List에 데이터 주입
                value_List.append(session_day_end_result)
                value_List.append(reconnect_count)

            else:
                # value_List에 데이터 주입
                value_List.append(session_day_end)
                value_List.append(reconnect_count)

            # 리스트에 초기화
            total_List.append(value_List)

        # 이중 배열에서 row를 담는 리스트 초기화
        value_List = []

        # updateDataList 리스트 초기화
        updateDataList =[]

        # total_List에 입력된 값 로그 테이블에 맞게 변경( 현재 날짜기준 startDate < date_now, session_endTime이 None인 경우)
        resultDataList=[]

        for i in range(len(total_List)):

            dataList = total_List[i]
            End_day = dataList[4]
            start_day = dataList[3]
            format = '%Y-%m-%d'

            date_now = datetime.datetime.now()
            date_now = date_now.strftime(format)
            date_now = datetime.datetime.strptime(date_now, format)

            start_day = start_day.strftime('%Y-%m-%d')
            start_day = datetime.datetime.strptime(start_day, format)

            if str(type(End_day)) != "<class 'NoneType'>":
                End_day = End_day.strftime('%Y-%m-%d')
                End_day = datetime.datetime.strptime(End_day, format)

            if ((str(type(End_day)) == "<class 'NoneType'>")) or (
                    (str(type(End_day)) != "<class 'NoneType'>") and start_day <= date_now):

                for j in range(len(dataList)):

                    if j == 3:
                        str_value = dataList[j]
                        format = '%Y-%m-%d %H:%M'
                        day_datetime = datetime.datetime.strftime(str_value, format)
                        updateDataList.append(day_datetime)

                    elif j == 4:
                        str_value = dataList[j]
                        if (str(type(str_value)) == "<class 'NoneType'>"):
                            str_value = "Running"
                            updateDataList.append(str_value)
                            str_value2 = "Calculating"
                            updateDataList.append(str_value2)

                        else:
                            format = '%Y-%m-%d %H:%M'
                            print(str_value)
                            day_datetime = datetime.datetime.strftime(str_value, format)
                            updateDataList.append(day_datetime)
                            if dataList[3] != "None" and dataList[4] != "None" and dataList[4] != "Running":
                                date_gap = dataList[4] - dataList[3]
                                gap_second = date_gap.seconds
                                gap_day = date_gap.days
                                gap_cal = date_gap.days * 24 * 60 * 60
                                total_gap = gap_cal + gap_second
                                str_value3 = (lambda x: '0' + x if len(x) < 8 else x)(
                                     str(datetime.timedelta(seconds=total_gap)))
                                print(gap_second)
                                print(str_value3)
                                updateDataList.append(str_value3)

                    elif j == 5:
                        str_value = dataList[j]
                        if (str(type(str_value)) == "<class 'NoneType'>"):
                            str_value = "0"
                            updateDataList.append(str_value)

                    else:
                        str_value = dataList[j]
                        updateDataList.append(str_value)

                resultDataList.append(updateDataList)
                updateDataList = []


        # 엑셀 문서가 저장될 폴더를 패스로 지정 (년도, 월이 바뀔 경우 폴더를 생성하고 년도나 월이 바뀔경우 해당 년도와 월로 생성)
        create_date_path = CreateDatePath()
        excel_path = create_date_path.create_folder()

        # 엑셀로 데이터 저장
        wb = openpyxl.Workbook()

        sheet1 = wb.active
        sheet1.title = "로그기록"
        sheetcolum = []
        sheetcolum.append("No")
        updatesearchList = []
        updateList = []

        for i in range(len(colum_list)):
            sheetcolum.append(colum_list[i])
        sheet1.append(sheetcolum)

        for i in range(len(resultDataList)):
            dataList = resultDataList[i]
            updateList.append((i + 1))
            print(updateList)
            for j in range(len(dataList)):
                value = dataList[j]
                updateList.append(value)
                print(updateList)
            updatesearchList.append(updateList)
            updateList = []

        for i in range(len(updatesearchList)):
            sheet1.append(updatesearchList[i])

        dt_now = datetime.datetime.now()
        current_time = dt_now.strftime('%Y-%m-%d %H:%M')
        print(current_time)

        str_titleName2 = str(current_time) + "분"
        current_time = current_time.replace("-", "년", 1)
        current_time = current_time.replace("-", "월")
        current_time = current_time.replace(" ", "일")
        current_time = current_time.replace(":", "시")
        current_time = current_time + "분"
        print(current_time)
        str_titleName = excel_path + '/(' + current_time + ')SaveLog.xlsx'

        max_row = sheet1.max_row + 1
        max_col = sheet1.max_column + 1

        for i in range(1, max_row):
            for j in range(1, max_col):
                cell = sheet1.cell(row=i, column=j)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(str_titleName)

        # win 라이브러리를 통해 엑셀 프로그램을 실행, 파일 불러옴
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')

            load_Wb = excel.Workbooks.Open(str_titleName)

            # 모든 엑셀 시트의 행과 열의 크기 자동 조절 진행
            for sh in load_Wb.Sheets:
                ws = load_Wb.Worksheets(sh.Name)
                ws.Columns.AutoFit()
                ws.Rows.AutoFit()

            # 수정한 엑셀 시트를 저장하고, 실행한 엑셀프로그램을 끝냄
            load_Wb.Save()
            excel.Application.Quit()
        except AttributeError:
            # Corner case dependencies.
            import os
            import re
            import sys
            import shutil
            # Remove cache and try again.
            MODULE_LIST = [m.__name__ for m in sys.modules.values()]
            for module in MODULE_LIST:
                if re.match(r'win32com\.gen_py\..+', module):
                    del sys.modules[module]
            shutil.rmtree(os.path.abspath(win32com.__gen_path__+'/..'))
            from win32com import client
            xl = client.gencache.EnsureDispatch('Excel.Application')

